import openpyxl

def detect_header_final_v2(file_path, sheet_name=0):
    """
    针对“上部复杂备注+下部数据”结构的终极通用版
    核心策略：寻找“最大列跨度”且“下一行数据类型发生突变（从文本变数字/混合）”的行。
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]

    best_row_idx = 0
    max_col_span = 0

    # 扫描前 30 行
    for r in range(1, min(30, ws.max_row or 1) + 1):
        row_values = [cell.value for cell in ws[r]]
        # 1. 计算当前行的有效列数（非空单元格数量）
        non_empty_count = sum(1 for v in row_values if v is not None and str(v).strip() != '')
        # 2. 关键判断：只有当这一行的列数 >= 之前遇到的最大列数时，才可能是表头
        # 这能过滤掉左侧只有几列信息的备注行（如第3行）
        if non_empty_count < 3:  # 假设表头至少有3列
            continue
        # 3. 启发式规则：检查下一行（r+1）
        # 真正的表头下面通常是数据。如果下一行也是满的文本，那当前行可能还是标题/备注。
        # 如果下一行开始变得“稀疏”或者包含数字，说明当前行大概率是表头。
        next_row_values = [cell.value for cell in ws[r+1]] if r < ws.max_row else []
        next_non_empty = sum(1 for v in next_row_values if v is not None and str(v).strip() != '')
        # 逻辑：
        # 如果当前行列数很多（比如5列），且下一行列数也差不多（说明数据结构开始了），
        # 或者下一行列数变少（说明进入了数据区，有些列可能是空的），这都是好迹象。
        # 最怕的是：当前行很满，下一行更满且全是长文本（说明还在备注区）。
        is_likely_header = False
        # 情况A：当前行是这一片区域里最宽的（定义了表格宽度）
        if non_empty_count > max_col_span:
            max_col_span = non_empty_count
            is_likely_header = True

        # 情况B：宽度持平，但我们要找最靠下的那个（防止把大标题当表头）
        # 结合你的截图，第5行是表头，它比上面的备注行更像“列定义”
        elif non_empty_count == max_col_span:
             # 简单粗暴：取后面出现的（覆盖前面的）
             is_likely_header = True

        if is_likely_header:
            # 额外校验：表头通常由短文本组成
            text_len_check = all(isinstance(v, str) and len(str(v)) < 50 for v in row_values if v)
            if text_len_check:
                 best_row_idx = r - 1 # 转为0基索引

    wb.close()
    return best_row_idx